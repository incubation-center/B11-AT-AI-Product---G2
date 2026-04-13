import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_test_case_excel(test_cases: list[dict]) -> io.BytesIO:
    """
    Generates a professional Excel file from a list of test case dictionaries.
    Expected dict keys: ID, Module, Title, Pre-conditions, Test Steps, Expected Result, Priority
    """
    df = pd.DataFrame(test_cases)
    
    # Define columns to ensure order even if AI misses some or adds others
    columns = ["ID", "Module", "Title", "Pre-conditions", "Test Steps", "Expected Result", "Priority"]
    # Reindex and handle missing columns
    df = df.reindex(columns=columns).fillna("-")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Test Cases")
        
        workbook = writer.book
        worksheet = writer.sheets["Test Cases"]
        
        # --- Styling ---
        # Header style: Dark blue background, white bold text
        header_fill = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Adjust column widths & cell alignment
        column_widths = {
            "ID": 10,
            "Module": 15,
            "Title": 30,
            "Pre-conditions": 35,
            "Test Steps": 45,
            "Expected Result": 35,
            "Priority": 12
        }

        for i, col in enumerate(columns):
            width = column_widths.get(col, 20)
            worksheet.column_dimensions[chr(65 + i)].width = width
            
            # Add alternating row colors and wrap text for data rows
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=i + 1)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border
                if row % 2 != 0: # Light gray for odd rows
                    cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    output.seek(0)
    return output
